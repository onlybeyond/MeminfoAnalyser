import openpyxl


def writeExcel(path, data_list, sheet_name=0, ):
    workbook = openpyxl.Workbook()
    writeToNewSheet(workbook, sheet_name, data_list)
    workbook.save(path)
    workbook.close()
    print("保存成功")


def writeToNewSheet(workbook, sheet_name, data_list):
    sheet = workbook[workbook.sheetnames[0]]
    if sheet_name != 0:
        sheet = workbook.create_sheet(sheet_name)
    index = len(data_list)
    for i in range(0, index):
        for j in range(0, len(data_list[i])):
            sheet.cell(row=i + 1, column=j + 1, value=data_list[i][j])
    return workbook
